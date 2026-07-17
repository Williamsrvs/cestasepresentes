from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, jsonify, redirect, url_for
from config import Config
import logging
import os
import mysql.connector as mysql_conn
from config import db_config
import openpyxl
from io import BytesIO
import xlsxwriter
from flask import send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from urllib.parse import quote

# Configuração de logging   
logging.basicConfig(
    level=getattr(logging, Config.LOG_LEVEL, logging.INFO),
    filename=Config.LOG_FILE,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

app = Flask(__name__)

# ✅ Carrega TODAS as configurações do Config automaticamente
app.config.from_object(Config)
app.config.update(
    SESSION_COOKIE_HTTPONLY=Config.SESSION_COOKIE_HTTPONLY,
    SESSION_COOKIE_SECURE=Config.SESSION_COOKIE_SECURE,
    SESSION_COOKIE_SAMESITE=Config.SESSION_COOKIE_SAMESITE,
    PERMANENT_SESSION_LIFETIME=Config.PERMANENT_SESSION_LIFETIME
)

# Filtro para converter bytes em base64 para exibir imagens
@app.template_filter('b64encode')
def b64encode_filter(data):
    if data is None:
        return ''
    import base64
    return base64.b64encode(data).decode('utf-8')

# Classe para gerenciar conexão com MySQL
class MySQLConnection:
    @staticmethod
    def get_connection():
        try:
            conn = mysql_conn.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password'],
                database=db_config['database'],
                port=db_config['port'],
                autocommit=False
            )
            return conn
        except mysql_conn.Error as e:
            logging.error(f"❌ Erro ao conectar ao MySQL: {e}")
            logging.error(f"   Host: {db_config['host']}")
            logging.error(f"   User: {db_config['user']}")
            logging.error(f"   Database: {db_config['database']}")
            raise Exception(f"Erro de conexão com banco de dados: {str(e)}")
        except Exception as e:
            logging.error(f"❌ Erro inesperado ao conectar: {e}")
            raise

# Alias para compatibilidade
mysql = MySQLConnection

# Log ao iniciar
logging.info("🚀 Aplicação iniciando...")
logging.info(f"📡 Conectando em: {Config.MYSQL_HOST}")
logging.info(f"👤 Usuário: {Config.MYSQL_USER}")
logging.info(f"🗄️  Database: {Config.MYSQL_DB}")

def ler_sql_robusto(schema_path):
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(schema_path, 'r', encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(schema_path, 'r', encoding='utf-8', errors='replace') as f:
        return f.read()


def criar_tabelas():
    try:
        with app.app_context():
            conn = MySQLConnection.get_connection()
            cur = conn.cursor(dictionary=True)
            schema_path = os.path.join(os.path.dirname(__file__), 'schema.sql')
            sql_text = ler_sql_robusto(schema_path)
            sql_commands = sql_text.split(';')
            for command in sql_commands:
                cmd = command.strip()
                if cmd and cmd.lower().startswith('create table'):
                    cur.execute(cmd)
            conn.commit()
            cur.close()
            conn.close()
            logging.info('✅ Tabelas criadas/verificadas com sucesso.')
    except Exception as e:
        logging.error(f'❌ Erro ao criar/verificar tabelas: {e}')

@app.route('/')
def index():
    conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
    try:
        # Buscar apenas produtos ativos (ativo = 1)   
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        produtos = cur.fetchall()
        
        # Buscar subgrupos únicos apenas de produtos ativos
        cur.execute("SELECT DISTINCT subgrupo FROM tbl_prod WHERE ativo = 1 ORDER BY subgrupo ASC")
        subgrupos = cur.fetchall()
        
        return render_template('index.html', produtos=produtos, subgrupos=subgrupos)
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos: {e}")
        return render_template('index.html', produtos=[], subgrupos=[])
    finally:
        cur.close()

@app.route('/cliente', methods=['GET', 'POST'])
def cliente():
    cur = None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_cliente")
            clientes = cur.fetchall()
            return render_template('cliente.html', clientes=clientes)
            
        elif request.method == 'POST':
            dados = request.form
            cur.execute("""
CREATE TABLE IF NOT EXISTS tbl_cliente (
    id_cliente INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    nome_cliente VARCHAR(100),
    telefone VARCHAR(50),
    email VARCHAR(50),
    endereco TEXT,
    bairro VARCHAR(50),
    cidade VARCHAR(50),
    uf VARCHAR(50),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")
            cur.execute("""
INSERT INTO tbl_cliente (nome_cliente, telefone, email, endereco, bairro, cidade, uf)
VALUES (%s,%s, %s, %s, %s, %s, %s)
""", (
    dados.get('nome_cliente'),
    dados.get('telefone'),
    dados.get('email'),
    dados.get('endereco'),
    dados.get('bairro'),
    dados.get('cidade'),
    dados.get('uf')
))
    
            conn.commit()
            logging.info(f"✅ Cliente cadastrado: {dados.get('nome_cliente')}")
            return redirect(url_for('cliente'))
            
    except Exception as e:
        logging.error(f"❌ Erro na rota /cliente: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

@app.route('/lojista', methods=['GET', 'POST'])
def lojista():
    return render_template('lojista.html')

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':
        nome_usuario = request.form.get('nome_usuario', '').strip()
        senha_usuario = request.form.get('senha_usuario', '').strip()

        # 🔹 LOGIN FIXO (ADMIN) — opcional
        if nome_usuario == 'admin' and senha_usuario == '220485':
            logging.info(f"✅ Login admin realizado: {nome_usuario}")
            return redirect(url_for('lojista'))  # ou lojista
        
        if nome_usuario == 'usuario_teste' and senha_usuario == 'teste123':
            logging.info(f"✅ Login admin realizado: {nome_usuario}")
            return redirect(url_for('lojista'))  # ou lojista

        try:
            conn = mysql.get_connection()
            cur = conn.cursor(dictionary=True)

            cur.execute("""
                SELECT id_usuario, nome_usuario 
                FROM tbl_cadastrar_usuario 
                WHERE nome_usuario = %s AND senha_usuario = %s
            """, (nome_usuario, senha_usuario))

            usuario = cur.fetchone()

            if usuario:
                logging.info(f"✅ Login bem-sucedido: {nome_usuario}")
                return redirect(url_for('lojista'))  # lojista.html
            else:
                logging.warning(f"❌ Usuário ou senha inválidos: {nome_usuario}")
                return render_template(
                    'login.html',
                    erro="Usuário ou senha inválidos!"
                )

        except Exception as e:
            logging.error(f"❌ Erro na rota /login: {e}")
            return render_template(
                'login.html',
                erro="Erro interno. Tente novamente."
            )

        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()

    return render_template('login.html')


@app.route('/cadastrar_usuario', methods=['GET', 'POST'])
def cadastrar_usuario():
    cur=None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_cadastrar_usuario")
            usuarios = cur.fetchall()
            return render_template('cadastrar_usuario.html', usuarios=usuarios)
        elif request.method == 'POST':
            dados = request.form
            # Validação de senha
            if dados.get('senha_usuario') != dados.get('repetir_senha'):
                return render_template('cadastrar_usuario.html', usuarios=[], erro="As senhas não conferem!")

            cur.execute("""
CREATE TABLE IF NOT EXISTS tbl_cadastrar_usuario(
    id_usuario INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    nome_usuario VARCHAR(100),
    senha_usuario VARCHAR(100),
    repetir_senha VARCHAR(100),
    dt_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")

            cur.execute("""
INSERT INTO tbl_cadastrar_usuario (nome_usuario, senha_usuario, repetir_senha)
VALUES (%s, %s, %s)
""", (
    dados.get('nome_usuario'),
    dados.get('senha_usuario'),
    dados.get('repetir_senha')
))
            conn.commit()
            logging.info(f"✅ Usuário cadastrado: {dados.get('nome_usuario')}")
            return redirect(url_for('cadastrar_usuario'))
    except Exception as e:
        logging.error(f"❌ Erro na rota /cadastrar_usuario: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"error": str(e)}),500
    finally:
        if cur:
            cur.close()


@app.route('/produto', methods=['GET', 'POST'])
@app.route('/produto/<int:id_prod>', methods=['PUT', 'DELETE'])
def produto(id_prod=None):
    cur = None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        
        # GET: Listar produtos (apenas ativos)
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
            produtos = cur.fetchall()
            return render_template('produto.html', produtos=produtos)
            
        # POST: Criar novo produto
        elif request.method == 'POST':
            dados = request.form
            imagem = request.files.get('imagem_url')
            
            # Ler os bytes da imagem se ela existir
            imagem_bytes = None
            if imagem and imagem.filename:
                imagem_bytes = imagem.read()
            
            # Criar tabela se não existir
            cur.execute("""CREATE TABLE IF NOT EXISTS tbl_prod (
                id_prod INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                nome_prod VARCHAR(100) NOT NULL,
                descricao TEXT,
                subgrupo varchar(50),
                status_promocao VARCHAR(50),
                valor DECIMAL(10,2) NOT NULL,
                form_pgmto VARCHAR(50),
                imagem_url MEDIUMBLOB,
                ativo TINYINT DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );""")
                
            cur.execute("""
                INSERT INTO tbl_prod (nome_prod, descricao, subgrupo, status_promocao, valor, form_pgmto, imagem_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                dados.get('nome_prod', ''),
                dados.get('descricao', ''),
                dados.get('subgrupo', ''),
                dados.get('status_promocao', ''),
                dados.get('valor', 0),
                dados.get('form_pgmto', ''),
                imagem_bytes
            ))
            conn.commit()
            logging.info(f"✅ Produto cadastrado: {dados.get('nome_prod')}")
            return redirect(url_for('produto'))
            
        # PUT: Atualizar produto existente
        elif request.method == 'PUT' and id_prod:
            dados = request.form
            imagem = request.files.get('imagem_url')
            
            if imagem and imagem.filename:
                imagem_bytes = imagem.read()
                cur.execute("""
                    UPDATE tbl_prod 
                    SET nome_prod = %s, descricao = %s, subgrupo = %s, status_promocao = %s, 
                        valor = %s, form_pgmto = %s, imagem_url = %s
                    WHERE id_prod = %s
                """, (
                    dados.get('nome_prod', ''),
                    dados.get('descricao', ''),
                    dados.get('subgrupo', ''),
                    dados.get('status_promocao', ''),
                    dados.get('valor', 0),
                    dados.get('form_pgmto', ''),
                    imagem_bytes,
                    id_prod
                ))
            else:
                cur.execute("""
                    UPDATE tbl_prod 
                    SET nome_prod = %s, descricao = %s, subgrupo = %s, status_promocao = %s, 
                        valor = %s, form_pgmto = %s
                    WHERE id_prod = %s
                """, (
                    dados.get('nome_prod', ''),
                    dados.get('descricao', ''),
                    dados.get('subgrupo', ''),
                    dados.get('status_promocao', ''),
                    dados.get('valor', 0),
                    dados.get('form_pgmto', ''),
                    id_prod
                ))
            
            conn.commit()
            logging.info(f"✅ Produto atualizado: {dados.get('nome_prod')}")
            return jsonify({"message": "Produto atualizado com sucesso"})

        # DELETE: Excluir produto (soft delete - marcar como inativo)
        elif request.method == 'DELETE' and id_prod:
            try:
                # Verificar se o produto existe e está ativo
                cur.execute("SELECT id_prod, nome_prod FROM tbl_prod WHERE id_prod = %s AND ativo = 1", (id_prod,))
                produto_existe = cur.fetchone()
                
                if not produto_existe:
                    return jsonify({"error": "Produto não encontrado ou já foi deletado"}), 404
                
                # Marcar como inativo (soft delete)
                cur.execute("UPDATE tbl_prod SET ativo = 0 WHERE id_prod = %s", (id_prod,))
                
                if cur.rowcount > 0:
                    conn.commit()
                    logging.info(f"✅ Produto desativado: {id_prod} - {produto_existe['nome_prod']}")
                    return jsonify({"message": "Produto excluído com sucesso"}), 200
                else:
                    conn.rollback()
                    return jsonify({"error": "Falha ao excluir o produto"}), 400
                    
            except Exception as delete_error:
                conn.rollback()
                logging.error(f"❌ Erro ao deletar produto {id_prod}: {delete_error}")
                return jsonify({"error": f"Erro ao excluir: {str(delete_error)}"}), 500
            
    except Exception as e:
        logging.error(f"❌ Erro na rota /produto: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

@app.route('/produto/<int:id_prod>/visibilidade', methods=['PATCH'])
def atualizar_visibilidade_produto(id_prod):
    """Atualizar visibilidade do produto (ativo/inativo)"""
    cur = None
    try:
        dados = request.get_json()
        
        if not dados or 'ativo' not in dados:
            return jsonify({"error": "Campo 'ativo' é obrigatório"}), 400
        
        status = 1 if dados.get('ativo') else 0
        
        conn = mysql.get_connection()
        cur = conn.cursor(dictionary=True)
        
        # Verificar se produto existe
        cur.execute("SELECT id_prod, nome_prod FROM tbl_prod WHERE id_prod = %s", (id_prod,))
        produto = cur.fetchone()
        
        if not produto:
            return jsonify({"error": "Produto não encontrado"}), 404
        
        # Atualizar visibilidade
        cur.execute("UPDATE tbl_prod SET ativo = %s WHERE id_prod = %s", (status, id_prod))
        conn.commit()
        
        status_text = "visível" if status == 1 else "oculto"
        logging.info(f"✅ Produto {id_prod} ({produto['nome_prod']}) agora está {status_text}")
        
        return jsonify({
            "message": f"Produto agora está {status_text}",
            "ativo": status
        }), 200
        
    except Exception as e:
        logging.error(f"❌ Erro ao atualizar visibilidade do produto {id_prod}: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

@app.route('/produto_excel', methods=['GET'])
def produto_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor (agora funcionando corretamente)
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        logging.info("✅ Cursor criado")

        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        produtos = cur.fetchall()

        logging.info(f"📊 Total de produtos: {len(produtos)}")

        if not produtos:
            return jsonify({"error": "Nenhum produto encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Produtos')

        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)

        logging.info("✍️ Cabeçalhos escritos")

        # --- LOOP CORRIGIDO ---
        for idx, produto in enumerate(produtos):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = produto[col]

                # Tratamentos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    # Evita texto ilegível de BLOB (imagem)
                    dados_linha.append("[BINARY_DATA]")
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(produtos)} linhas escritas")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'produtos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()


@app.route('/teste_produtos', methods=['GET'])
def teste_produtos():
    cur = None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC LIMIT 3")

        colunas = [i[0] for i in cur.description]
        produtos = cur.fetchall()

        return jsonify({
            "status": "success",
            "total": len(produtos),
            "colunas": colunas,
            "produtos": produtos
        })

    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "erro": str(e),
            "traceback": traceback.format_exc()
        }), 500

    finally:
        if cur:
            cur.close()
            
@app.route('/pesquisa', methods=['GET', 'POST'])
def pesquisa():
    print(f"\n{'='*50}")
    print(f"MÉTODO: {request.method}")
    print(f"{'='*50}\n")
    
    if request.method == "POST":
        print("✅ Entrou no POST")
        print(f"📝 Dados recebidos: {dict(request.form)}\n")
        
        try:
            print("🔌 Tentando conectar ao MySQL...")
            conn = mysql.get_connection()
            cur = conn.cursor(dictionary=True)
            print("✅ Conectado ao banco!\n")

            # Criação da tabela
            print("📋 Criando/verificando tabela...")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tbl_pesquisa_satisfacao (
                    id_pesquisa INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                    atendimento VARCHAR(50),
                    qualidade VARCHAR(50),
                    satisfacao VARCHAR(50),
                    rapidez VARCHAR(50),
                    localizacao VARCHAR(50),
                    experiencia VARCHAR(50),
                    facilidade VARCHAR(50),
                    variedade VARCHAR(50),
                    ambiente VARCHAR(50),
                    recomendacao VARCHAR(50),
                    comentarios TEXT,
                    data_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            print("✅ Tabela OK!\n")

            # Coleta dos dados
            campos = ['atendimento', 'qualidade', 'satisfacao', 'rapidez', 'localizacao',
                    'experiencia', 'facilidade', 'variedade', 'ambiente', 'recomendacao']
            
            valores = [request.form.get(c) for c in campos]
            comentarios = request.form.get('comentarios', '')
            
            print("📊 Valores a serem inseridos:")
            for campo, valor in zip(campos, valores):
                print(f"  - {campo}: {valor}")
            print(f"  - comentarios: {comentarios}\n")

            # Inserção no banco
            print("💾 Inserindo no banco...")
            cur.execute("""
                INSERT INTO tbl_pesquisa_satisfacao (
                    atendimento, qualidade, satisfacao, rapidez, localizacao,
                    experiencia, facilidade, variedade, ambiente, recomendacao, comentarios
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (*valores, comentarios))
            
            print(f"✅ Linhas afetadas: {cur.rowcount}")
            
            conn.commit()
            print("✅ COMMIT realizado com sucesso!")
            
            cur.close()
            conn.close()
            print("🔒 Conexão fechada\n")

            return render_template('pesquisa.html', mensagem='Pesquisa enviada com sucesso!')

        except Exception as e:
            print(f"\n❌ ERRO CAPTURADO: {str(e)}\n")
            import traceback
            traceback.print_exc()
            return render_template('pesquisa.html', erro=str(e))
    
    print("📄 Renderizando formulário (GET)\n")
    return render_template('pesquisa.html')

@app.route('/contato', methods=['GET', 'POST'])
def contato():
    if request.method == 'POST':
        try:
            nome = request.form.get('nome')
            email = request.form.get('email')
            mensagem = request.form.get('mensagem')
            
            # Validação básica
            if not nome or not email or not mensagem:
                return render_template('contato.html', 
                                    erro="Todos os campos são obrigatórios!")
            
            # Criar cursor
            conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
            
            # Inserir dados
            cur.execute("""
                INSERT INTO u799109175_cestas_present.tbl_fale_conosco 
                (nome, email, mensagem)
                VALUES (%s, %s, %s)
            """, (nome, email, mensagem))
            
            # ORDEM CORRETA: commit antes de fechar
            conn.commit()
            cur.close()
            
            logging.info(f"📩 Mensagem recebida de {nome} ({email})")
            
            return render_template('contato.html', 
                                sucesso="Mensagem enviada com sucesso!")
        
        except Exception as e:
            logging.error(f"❌ Erro ao processar contato: {e}")
            return render_template('contato.html', 
                                erro="Erro ao enviar mensagem. Tente novamente.")
    
    return render_template('contato.html')
if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 SERVICE TOUR - Iniciando servidor...")
    print("="*50)

    import socket
    local_ip = socket.gethostbyname(socket.gethostname())

    print(f"📍 Servidor local: http://127.0.0.1:5000")
    print(f"📍 Servidor LAN: http://{local_ip}:5000")
    print("="*50 + "\n")

    try:
        # Testa conexão antes de iniciar
        with app.app_context():
            conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
            cur.execute("SELECT DATABASE(), VERSION()")
            db, version = cur.fetchone()
            cur.close()
            print(f"✅ Conectado ao banco: {db}")
            print(f"✅ Versão MySQL: {version}\n")
        # Chama função para criar/verificar tabelas
        criar_tabelas()
    except Exception as e:
        print(f"⚠️  Aviso: Não foi possível conectar ao banco: {e}\n")


from datetime import datetime

@app.route('/ger_clientes', methods=['GET', 'POST'])
def ger_clientes():
    if request.method == 'GET':
        try:
            conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
            
            # Captura os parâmetros de data
            data_inicio = request.args.get('data_inicio')
            data_fim = request.args.get('data_fim')
            
            # Query base 
            query = "SELECT * FROM tbl_cliente"
            params = []
            
            # Adiciona filtros se as datas forem fornecidas
            if data_inicio and data_fim:
                query += " WHERE DATE(created_at) BETWEEN %s AND %s"
                params = [data_inicio, data_fim]
            elif data_inicio:
                query += " WHERE DATE(created_at) >= %s"
                params = [data_inicio]
            elif data_fim:
                query += " WHERE DATE(created_at) <= %s"
                params = [data_fim]
            
            query += " ORDER BY nome_cliente ASC"
            
            # Executa a query com ou sem parâmetros
            if params:
                cur.execute(query, params)
            else:
                cur.execute(query)
                
            clientes = cur.fetchall()
            cur.close()
            
            return render_template('ger_clientes.html', clientes=clientes)
            
        except Exception as e:
            logging.error(f"❌ Erro ao buscar clientes: {e}")
            return render_template('ger_clientes.html', clientes=[])
    
    return render_template('ger_clientes.html')


@app.route('/cliente_excel', methods=['GET'])
def cliente_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM tbl_cliente")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Cliente: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum cliente encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Clientes')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


@app.route('/pedidos', methods=['GET', 'POST'])
def pedidos():
    cur = None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        produtos = cur.fetchall()
        cur.execute("SELECT * FROM tbl_cliente")
        clientes = cur.fetchall()
        return render_template('pedidos.html', produtos=produtos, clientes=clientes)
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos para pedidos: {e}")
        return render_template('pedidos.html', produtos=[], clientes=[])
    finally:
        if cur:
            cur.close()


@app.route('/api/produtos', methods=['GET'])
def api_produtos():
    """Endpoint para carregar produtos via AJAX"""
    cur = None
    try:
        conn = mysql.get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id_prod, nome_prod, valor FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        produtos = cur.fetchall()
        
        if not produtos:
            return jsonify({"status": "sucesso", "produtos": [], "mensagem": "Nenhum produto encontrado"})
        
        return jsonify({
            "status": "sucesso",
            "produtos": produtos,
            "total": len(produtos)
        })
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos via API: {e}")
        return jsonify({
            "status": "erro",
            "mensagem": str(e),
            "produtos": []
        }), 500
    finally:
        if cur:
            cur.close()


@app.route('/salvar_pedido', methods=['POST'])
def salvar_pedido():
    cur = None
    try:
        dados = request.get_json()
        
        if not dados or 'carrinho' not in dados:
            return jsonify({"status": "erro", "mensagem": "Dados inválidos"}), 400
        
        carrinho = dados.get('carrinho', [])
        id_cliente = dados.get('id_cliente')
        nome_cliente = dados.get('nome_cliente')
        # O frontend envia como email_cliente e telefone_cliente, mas o BD usa email e telefone
        email = dados.get('email_cliente', dados.get('email', ''))
        telefone = dados.get('telefone_cliente', dados.get('telefone', ''))
        numero_mesa = dados.get('numero_mesa') or 0  # Se vazio/null, usar 0
        endereco = dados.get('endereco', '')
        bairro = dados.get('bairro', '')
        cidade = dados.get('cidade', '')
        uf = dados.get('uf', '')
        ponto_referencia = dados.get('ponto_referencia', '')
        form_pgmto = dados.get('form_pgmto', '')
        tipo_consumo = dados.get('tipo_consumo', '')
        observacao = dados.get('observacao', '')
        taxa_entrega = dados.get('taxa_entrega', 0.0)

        if not carrinho:
            return jsonify({"status": "erro", "mensagem": "Carrinho vazio"}), 400

        # Abrir conexão o mais cedo possível para permitir criar/consultar cliente
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)

        # Se não enviaram id_cliente, tentar encontrar por telefone/nome ou criar novo cliente
        if not id_cliente:
            if nome_cliente and nome_cliente.strip() != '':
                # Tentar correspondência por telefone primeiro quando disponível
                encontrado = None
                try:
                    if telefone and telefone.strip() != '':
                        cur.execute("SELECT id_cliente FROM tbl_cliente WHERE telefone = %s LIMIT 1", (telefone,))
                        encontrado = cur.fetchone()
                    # Se não encontrou por telefone, tentar por nome
                    if not encontrado:
                        cur.execute("SELECT id_cliente FROM tbl_cliente WHERE nome_cliente = %s LIMIT 1", (nome_cliente,))
                        encontrado = cur.fetchone()

                    if encontrado and encontrado.get('id_cliente'):
                        id_cliente = encontrado.get('id_cliente')
                    else:
                        # Inserir novo cliente com TODOS os dados fornecidos
                        cur.execute(
                            "INSERT INTO tbl_cliente (nome_cliente,email,endereco,bairro,cidade,uf,telefone) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                            (
                                nome_cliente,
                                email,      # Email do formulário
                                endereco or '',
                                bairro or '',
                                cidade or '',       # Cidade do formulário
                                uf or '',           # UF do formulário
                                telefone
                            )
                        )
                        id_cliente = cur.lastrowid
                        logging.info(f"✅ Cliente criado: {id_cliente} - {nome_cliente} | Email: {email} | Cidade: {cidade}, {uf}")
                except Exception as e:
                    logging.error(f"❌ Erro ao localizar/criar cliente: {e}")
                    conn.rollback()
                    return jsonify({"status": "erro", "mensagem": "Erro ao processar cliente"}), 500
            else:
                # Nem id nem nome foram informados
                cur.close()
                return jsonify({"status": "erro", "mensagem": "ID ou nome do cliente obrigatório"}), 400

        # Calcular valor total do pedido
        valor_total = sum(float(item.get('subtotal', 0)) for item in carrinho)
        
        # Inserir pedido principal
        cur.execute("""
            INSERT INTO tbl_pedidos (id_cliente, valor_total, numero_mesa)
            VALUES (%s, %s, %s)
        """, (id_cliente, valor_total, numero_mesa))
        
        id_pedido = cur.lastrowid
        logging.info(f"✅ Pedido criado: {id_pedido}")
        
        # Inserir detalhes do pedido para cada item no carrinho
        for item in carrinho:
            id_prod = item.get('produtoId')
            quantidade = item.get('quantidade')
            preco_unitario = item.get('valor')
            valor_item = float(item.get('subtotal', 0))  # Subtotal do item
            
            cur.execute("""
                INSERT INTO tbl_detalhes_pedido 
                (id_pedido, id_prod, id_cliente, quantidade, preco_unitario, nome_cliente, telefone, valor_total, numero_mesa, endereco, bairro, ponto_referencia, form_pgmto, tipo_consumo, observacao,taxa_entrega)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (id_pedido, id_prod, id_cliente, quantidade, preco_unitario, nome_cliente, telefone, valor_item, numero_mesa, endereco, bairro, ponto_referencia, form_pgmto, tipo_consumo, observacao,taxa_entrega))
        
        conn.commit()
        logging.info(f"✅ Pedido salvo: {id_pedido} com {len(carrinho)} itens")
        
        return jsonify({
            "status": "sucesso",
            "mensagem": "Pedido salvo com sucesso!",
            "id_pedido": id_pedido,
            "valor_total": float(valor_total),
            "taxa_entrega": float(taxa_entrega) if taxa_entrega else 0.0
        }), 200

        
    except Exception as e:
        logging.error(f"❌ Erro ao salvar pedido: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"status": "erro", "mensagem": str(e)}), 500
    finally:
        if cur:
            cur.close()

# Criar uma rota para enviar os pedidos via whatsapp
@app.route('/enviar_whatsapp', methods=['POST'])
def enviar_whatsapp():
    """
    Gera link wa.me para enviar pedido via WhatsApp 
    NÃO salva pedido - apenas gera e retorna URL
    """
    try:
        dados = request.get_json()
        
        if not dados:
            return jsonify({
                "status": "erro",
                "mensagem": "Dados não fornecidos"
            }), 400
        
        whatsapp_numero = dados.get('whatsapp_numero')
        mensagem = dados.get('mensagem')
        
        if not whatsapp_numero or not mensagem:
            logging.warning(f"⚠️ Parâmetros incompletos: numero={whatsapp_numero}, msg_length={len(str(mensagem)) if mensagem else 0}")
            return jsonify({
                "status": "erro",
                "mensagem": "Número de WhatsApp ou mensagem não fornecidos"
            }), 400
        
        # Limpar número (remover caracteres especiais)
        whatsapp_numero_limpo = ''.join(filter(str.isdigit, str(whatsapp_numero)))
        
        if not whatsapp_numero_limpo:
            logging.warning(f"❌ Número WhatsApp inválido após limpeza: {whatsapp_numero}")
            return jsonify({
                "status": "erro",
                "mensagem": "Número de WhatsApp inválido"
            }), 400
        
        try:
            # Criar URL do WhatsApp com mensagem pré-formatada
            # quote() codifica a mensagem para ser segura em URL
            url_whatsapp = f"https://wa.me/{whatsapp_numero_limpo}?text={quote(str(mensagem))}"
            
            logging.info(f"📱 Link WhatsApp gerado com sucesso para {whatsapp_numero_limpo}")
            
            return jsonify({
                "status": "sucesso",
                "mensagem": "Link WhatsApp gerado com sucesso!",
                "numero_whatsapp": whatsapp_numero_limpo,
                "url_whatsapp": url_whatsapp
            }), 200
        
        except Exception as e:
            logging.error(f"❌ Erro ao gerar URL WhatsApp: {e}")
            return jsonify({
                "status": "erro",
                "mensagem": f"Erro ao processar mensagem: {str(e)}"
            }), 500
    
    except Exception as e:
        logging.error(f"❌ Erro na rota enviar_whatsapp: {e}", exc_info=True)
        return jsonify({
            "status": "erro",
            "mensagem": f"Erro ao processar requisição: {str(e)}"
        }), 500


@app.route('/gerar_brcode_pix', methods=['POST'])
def gerar_brcode_pix():
    """
    Gera QR Code PIX válido (Brcode) do Banco Central do Brasil
    
    Recebe:
    - chave_pix: CPF, email, telefone ou chave aleatória
    - valor: Valor em reais (opcional)
    - nome_beneficiario: Nome do beneficiário
    
    Retorna: QR Code em base64 para exibição na página
    """
    try:
        from brcode import BRCode
        import qrcode
        import io
        import base64
        
        dados = request.get_json()
        chave_pix = dados.get('chave_pix', '')
        valor = float(dados.get('valor', 0))
        nome_beneficiario = dados.get('nome_beneficiario', 'LOJA')
        
        if not chave_pix:
            return jsonify({
                "status": "erro",
                "mensagem": "Chave PIX não fornecida"
            }), 400
        
        # Gerar Brcode PIX válido (texto cópia e cola)
        pix_dict = {
            'name': nome_beneficiario,
            'city': 'Brasilia',  # Cidade padrão
            'account': 'br.gov.bcb.brcode',
            'key': chave_pix
        }
        
        if valor > 0:
            pix_dict['amount'] = valor
        
        # Usar biblioteca brcode para gerar string Brcode
        brcode = BRCode(**pix_dict)
        brcode_texto = str(brcode)
        
        # Gerar QR Code a partir do Brcode texto
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(brcode_texto)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Converter para base64
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        logging.info(f"✅ Brcode PIX gerado com sucesso: {chave_pix[:10]}... | Valor: R$ {valor}")
        
        return jsonify({
            "status": "sucesso",
            "qr_code_base64": f"data:image/png;base64,{img_base64}",
            "brcode_texto": brcode_texto,
            "chave_pix": chave_pix,
            "valor": valor
        }), 200
        
    except Exception as e:
        logging.error(f"❌ Erro ao gerar Brcode PIX: {e}", exc_info=True)
        return jsonify({
            "status": "erro",
            "mensagem": f"Erro ao gerar QR Code PIX: {str(e)}"
        }), 500


@app.route('/ger_pedidos', methods=['GET', 'POST'])
def ger_pedidos():
    if request.method == 'GET':
        try:
            conn = mysql.get_connection()
            cur = conn.cursor(dictionary=True)

            data_inicio = request.args.get('data_inicio').strftime('%d/%m/%Y') if request.args.get('data_inicio') else None
            data_fim = request.args.get('data_fim').strftime('%d/%m/%Y') if request.args.get('data_fim') else None

            if data_inicio:
                data_inicio = datetime.strptime(data_inicio,'%Y-%m-%d').strftime('%d/%m/%Y')
            if data_fim:
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')


            query = """SELECT * FROM vw_pedidos_fin
                    where date(dt_registro) <= curdate()
                    
            """  
            params = []

            if data_inicio and data_fim:
                query += " WHERE DATE(dt_registro) BETWEEN %s AND %s"
                params = [data_inicio, data_fim]
            elif data_inicio:
                
                query += " WHERE DATE(dt_registro) >= %s"
                params = [data_inicio]
            elif data_fim:
                query += " WHERE DATE(dt_registro) <= %s" 
                params = [data_fim]

            query += " ORDER BY dt_registro DESC"

            if params:
                cur.execute(query, params)
            else:
                cur.execute(query)

            pedidos = cur.fetchall()
            cur.close()

            total_valor = sum([p['valor_total'] for p in pedidos])
            qtde_total = len(set([p['id_pedido'] for p in pedidos]))
            qtde_itens = sum([p['qtde'] for p in pedidos])
            

            return render_template(
                'ger_pedidos.html',
                pedidos=pedidos,
                total_valor=total_valor,
                qtde_total=qtde_total, 
                qtde_itens=qtde_itens,
                
            )

        except Exception as e:
            logging.error(f"❌ Erro ao buscar pedidos: {e}")
            return render_template(
                'ger_pedidos.html',
                pedidos=[],
                total_valor=0,     
                qtde_total=0,
                qtde_itens=0,
                
            )

    return render_template('ger_pedidos.html', 
    pedidos=[], total_valor=0, qtde_total=0, qtde_itens=0)


@app.route('/pedidos_excel', methods=['GET'])
def pedidos_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM vw_pedidos_fin")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Pedidos: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum pedido encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Pedidos')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


#=============================================================

@app.route('/pedidos_excel_clientes', methods=['GET'])
def pedidos_excel_clientes():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM vw_resumo_pedidos_cliente")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Pedidos: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum pedido encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Pedidos')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pedidos_cliente_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


@app.route('/nav_tabs', methods=['GET'])
def nav_tabs():
    try:
        conn = mysql.get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT DISTINCT subgrupo FROM tbl_prod ORDER BY subgrupo ASC")
        subgrupos = cur.fetchall()

    except Exception as e:
        logging.error(f"❌ Erro ao buscar subgrupos: {e}")
        subgrupos = []

    finally:
        if cur:
            cur.close()

    return render_template('nav_tabs.html', subgrupos=subgrupos)    

@app.route('/subgrupo/<string:subgrupo>', methods=['GET'])
def grupo(subgrupo):
    try:
        conn = mysql.get_connection()
        cur = conn.cursor(dictionary=True)
        
        # Buscar produtos do subgrupo selecionado
        cur.execute("SELECT * FROM tbl_prod WHERE subgrupo = %s AND ativo = 1 ORDER BY created_at DESC", (subgrupo,))
        produtos = cur.fetchall()
        
        # Buscar todos os subgrupos para o filtro
        cur.execute("SELECT DISTINCT subgrupo FROM tbl_prod WHERE ativo = 1 ORDER BY subgrupo ASC")
        subgrupos = cur.fetchall()

    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos do grupo: {e}")
        produtos = []
        subgrupos = []

    finally:
        if cur:
            cur.close()

    return render_template('index.html', produtos=produtos, subgrupos=subgrupos, subgrupo_selecionado=subgrupo)

@app.route('/pedidos_cliente', methods=['GET', 'POST'])
def pedidos_cliente():
    cur = None
    try:
        conn = mysql.get_connection()
        cur = conn.cursor(dictionary=True)
        
        # ✅ Buscar apenas as colunas necessárias para o JavaScript
        cur.execute("SELECT id_prod, nome_prod, valor FROM tbl_prod WHERE ativo = 1 ORDER BY nome_prod ASC")
        produtos = cur.fetchall()
        
        # ✅ Buscar clientes
        cur.execute("SELECT id_cliente, nome_cliente FROM tbl_cliente ORDER BY nome_cliente ASC")
        clientes = cur.fetchall()
        
        print(f"✅ Produtos carregados: {len(produtos)}")
        print(f"📦 Primeiro produto: {produtos[0] if produtos else 'Nenhum'}")
        
        return render_template('pedidos_cliente.html', produtos=produtos, clientes=clientes)
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos para pedidos: {e}")
        print(f"❌ ERRO: {e}")
        return render_template('pedidos_cliente.html', produtos=[], clientes=[])
    finally:
        if cur:
            cur.close()


@app.route('/imprimirSelecionados', methods=['POST'])
def imprimirSelecionados():
    ids = request.json.get('pedidos')

    # Buscar pedidos no banco
    pedidos = buscar_pedidos(ids)

    return render_template('imprimirSelecionados.html', pedidos=pedidos)


@app.route('/entregadores', methods=['GET', 'POST'])
def entregadores():
    cur = None
    try:
        conn = mysql.get_connection(); cur = conn.cursor(dictionary=True)
        
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_entregadores")
            entregadores = cur.fetchall()
            return render_template('entregadores.html', entregadores=entregadores)

        elif request.method == 'POST':
            dados = request.form
            cur.execute("""
CREATE TABLE IF NOT EXISTS tbl_entregadores (
    id_entregador INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    nome_entregador VARCHAR(100),
    habilitacao VARCHAR(50),
    tipo_cnh enum('A', 'B', 'C', 'D', 'E', 'AB','ACC') NOT NULL,
    validade_cnh DATE,
    endereco VARCHAR(200),
    bairro VARCHAR(50),
    cidade VARCHAR(50),
    uf VARCHAR(50),
    telefone VARCHAR(50),
    veiculo VARCHAR(50),
    ano_veiculo INT,
    cor VARCHAR(50),
    placa VARCHAR(20),
    ativo TINYINT DEFAULT 1,
    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")
            cur.execute("""
INSERT INTO tbl_entregadores (nome_entregador, habilitacao, tipo_cnh, validade_cnh, endereco, bairro, cidade, uf, telefone, veiculo, ano_veiculo, cor, placa)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
""", (
    dados.get('nome_entregador'),
    dados.get('habilitacao'),
    dados.get('tipo_cnh'),
    dados.get('validade_cnh'),
    dados.get('endereco'),
    dados.get('bairro'),
    dados.get('cidade'),
    dados.get('uf'),
    dados.get('telefone'),
    dados.get('veiculo'),
    dados.get('ano_veiculo'),
    dados.get('cor'),
    dados.get('placa')
))
    
            conn.commit()
            logging.info(f"✅ Entregador cadastrado: {dados.get('nome_entregador')}")
            return redirect(url_for('entregadores'))

    except Exception as e:
        logging.error(f"❌ Erro na rota /entregadores: {e}")
        if conn is not None:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

# Criar uma rota para dar baixa nos pedidos enviados.

# Substitua sua rota atual por esta:
@app.route('/atualizar_status_pedido', methods=['POST'])
def atualizar_status_pedido():
    try:
        dados       = request.get_json(force=True, silent=True)
        id_pedido   = dados.get('id_pedido')
        novo_status = dados.get('status')

        status_validos = ['Pendente', 'Em producao', 'Cancelado', 'Finalizado']
        if novo_status not in status_validos:
            return jsonify({'sucesso': False, 'erro': 'Status inválido'}), 400

        conn = mysql.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "UPDATE tbl_detalhes_pedido SET status_pedido = %s WHERE id_pedido = %s",
            (novo_status, id_pedido)
        )
        conn.commit()
        cur.close()

        return jsonify({'sucesso': True})

    except Exception as e:
        logging.error(f"❌ Erro ao atualizar status: {e}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/dashboard',methods=['GET'])
def dashboard():

    return render_template('dashboard.html')

