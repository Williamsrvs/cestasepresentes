from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, redirect, url_for
from config import Config
import logging
import os
import threading
import mysql.connector as mysql_conn
import openpyxl
from io import BytesIO
import xlsxwriter
from flask import send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from urllib.parse import quote

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    filename='app_errors.log',
    format='%(asctime)s - %(levelname)s - %(message)s'
)

app = Flask(__name__)

# ✅ Carrega TODAS as configurações do Config automaticamente
app.config.from_object(Config)

class MySQL:
    def __init__(self, app=None):
        self.app = app
        self._local = threading.local()

    def _new_connection(self):
        try:
            conn = mysql_conn.connect(
                host=Config.MYSQL_HOST,
                user=Config.MYSQL_USER,
                password=Config.MYSQL_PASSWORD,
                database=Config.MYSQL_DB,
                port=Config.MYSQL_PORT,
                autocommit=False
            )
            return conn
        except mysql_conn.Error as e:
            logging.error(f"❌ Falha ao conectar ao MySQL: {e}")
            raise

    @property
    def connection(self):
        conn = getattr(self._local, 'connection', None)
        if conn is None or not conn.is_connected():
            conn = self._new_connection()
            self._local.connection = conn
        return conn

    def commit(self):
        try:
            self.connection.commit()
        except Exception as e:
            logging.error(f"❌ Falha ao commitar no MySQL: {e}")
            raise

    def rollback(self):
        conn = getattr(self._local, 'connection', None)
        if conn is not None and conn.is_connected():
            try:
                conn.rollback()
            except Exception as e:
                logging.error(f"❌ Falha ao rollback no MySQL: {e}")
                raise

    def close(self):
        conn = getattr(self._local, 'connection', None)
        if conn is not None:
            try:
                conn.close()
            except Exception as e:
                logging.error(f"❌ Falha ao fechar conexão MySQL: {e}")
            finally:
                self._local.connection = None

mysql = MySQL(app)

# Filtro para converter bytes em base64 para exibir imagens
@app.template_filter('b64encode')
def b64encode_filter(data):
    if data is None:
        return ''
    import base64
    return base64.b64encode(data).decode('utf-8')

logging.info(f"👤 Usuário: {Config.MYSQL_USER}")
logging.info(f"🗄️  Database: {Config.MYSQL_DB}")

def ler_sql_robusto(schema_path):
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(schema_path, 'r', encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(schema_path, 'r', encoding='utf-8', errors='replace') as f:
        return f.read()


def criar_tabelas():
    try:
        with app.app_context():
            cur = mysql.connection.cursor()
            schema_path = os.path.join(os.path.dirname(__file__), 'schema.sql')
            sql_text = ler_sql_robusto(schema_path)
            sql_commands = sql_text.split(';')
            for command in sql_commands:
                cmd = command.strip()
                if cmd and cmd.lower().startswith('create table'):
                    cur.execute(cmd)
            mysql.connection.commit()
            cur.close()
            logging.info('✅ Tabelas criadas/verificadas com sucesso.')
    except Exception as e:
        logging.error(f'❌ Erro ao criar/verificar tabelas: {e}')

@app.route('/')
def index():
    cur = mysql.connection.cursor()
    try:
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        produtos = cur.fetchall()
        
        cur.execute("SELECT DISTINCT subgrupo FROM tbl_prod ORDER BY subgrupo ASC")
        subgrupos_raw = cur.fetchall()
        subgrupos = [{"subgrupo": s[0]} for s in subgrupos_raw] if subgrupos_raw else []
        
        return render_template('index.html', produtos=produtos, subgrupos=subgrupos)
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos: {e}")
        return render_template('index.html', produtos=[], subgrupos=[])
    finally:
        cur.close()

@app.route('/cliente', methods=['GET', 'POST'])
def cliente():
    cur = None
    try:
        cur = mysql.connection.cursor()
        
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_cliente")
            clientes = cur.fetchall()
            return render_template('cliente.html', clientes=clientes)
            
        elif request.method == 'POST':
            dados = request.form
            cur.execute("""
CREATE TABLE IF NOT EXISTS tbl_cliente (
    id_cliente INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    nome_cliente VARCHAR(100),
    telefone VARCHAR(50),
    email VARCHAR(50),
    endereco TEXT,
    bairro VARCHAR(50),
    cidade VARCHAR(50),
    uf VARCHAR(50),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")
            cur.execute("""
INSERT INTO tbl_cliente (nome_cliente, telefone, email, endereco, bairro, cidade, uf)
VALUES (%s,%s, %s, %s, %s, %s, %s)
""", (
    dados.get('nome_cliente'),
    dados.get('telefone'),
    dados.get('email'),
    dados.get('endereco'),
    dados.get('bairro'),
    dados.get('cidade'),
    dados.get('uf')
))
    
            mysql.connection.commit()
            logging.info(f"✅ Cliente cadastrado: {dados.get('nome_cliente')}")
            return redirect(url_for('cliente'))
            
    except Exception as e:
        logging.error(f"❌ Erro na rota /cliente: {e}")
        if mysql.connection:
            mysql.connection.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

@app.route('/lojista', methods=['GET', 'POST'])
def lojista():
    return render_template('lojista.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        nome_usuario = request.form.get('nome_usuario')
        senha_usuario = request.form.get('senha_usuario')
        if nome_usuario == 'admin' and senha_usuario == '859117':
            success_message = f"✅ Login bem-sucedido para o usuário: {nome_usuario}"
            return render_template('lojista.html', success_message=success_message)
        else:
            logging.warning(f"❌ Login falhou para o usuário: {nome_usuario}")
            return render_template('login.html', erro="Usuário ou senha inválidos!")
    return render_template('login.html')

@app.route('/cadastrar_usuario', methods=['GET', 'POST'])
def cadastrar_usuario():
    cur=None
    try:
        cur = mysql.connection.cursor()
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_cadastrar_usuario")
            usuarios = cur.fetchall()
            return render_template('cadastrar_usuario.html', usuarios=usuarios)
        elif request.method == 'POST':
            dados = request.form
            # Validação de senha
            if dados.get('senha_usuario') != dados.get('repetir_senha'):
                return render_template('cadastrar_usuario.html', usuarios=[], erro="As senhas não conferem!")

            cur.execute("""
CREATE TABLE IF NOT EXISTS tbl_cadastrar_usuario(
    id_usuario INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    nome_usuario VARCHAR(100),
    senha_usuario VARCHAR(100),
    repetir_senha VARCHAR(100),
    dt_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")

            cur.execute("""
INSERT INTO tbl_cadastrar_usuario (nome_usuario, senha_usuario, repetir_senha)
VALUES (%s, %s, %s)
""", (
    dados.get('nome_usuario'),
    dados.get('senha_usuario'),
    dados.get('repetir_senha')
))
            mysql.connection.commit()
            logging.info(f"✅ Usuário cadastrado: {dados.get('nome_usuario')}")
            return redirect(url_for('cadastrar_usuario'))
    except Exception as e:
        logging.error(f"❌ Erro na rota /cadastrar_usuario: {e}")
        if mysql.connection:
            mysql.connection.rollback()
        return jsonify({"error": str(e)}),500
    finally:
        if cur:
            cur.close()


@app.route('/produto', methods=['GET', 'POST'])
@app.route('/produto/<int:id_prod>', methods=['PUT', 'DELETE'])
def produto(id_prod=None):
    cur = None
    try:
        cur = mysql.connection.cursor()
        
        # GET: Listar produtos (apenas ativos)
        if request.method == 'GET':
            cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
            produtos = cur.fetchall()
            return render_template('produto.html', produtos=produtos)
            
        # POST: Criar novo produto
        elif request.method == 'POST':
            dados = request.form
            imagem = request.files.get('imagem_url')
            
            # Ler os bytes da imagem se ela existir
            imagem_bytes = None
            if imagem and imagem.filename:
                imagem_bytes = imagem.read()
            
            # Criar tabela se não existir
            cur.execute("""CREATE TABLE IF NOT EXISTS tbl_prod (
                id_prod INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                nome_prod VARCHAR(100) NOT NULL,
                descricao TEXT,
                subgrupo varchar(50),
                status_promocao VARCHAR(50),
                valor DECIMAL(10,2) NOT NULL,
                form_pgmto VARCHAR(50),
                imagem_url MEDIUMBLOB,
                ativo TINYINT DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );""")
                
            cur.execute("""
                INSERT INTO tbl_prod (nome_prod, descricao, subgrupo, status_promocao, valor, form_pgmto, imagem_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                dados.get('nome_prod', ''),
                dados.get('descricao', ''),
                dados.get('subgrupo', ''),
                dados.get('status_promocao', ''),
                dados.get('valor', 0),
                dados.get('form_pgmto', ''),
                imagem_bytes
            ))
            mysql.connection.commit()
            logging.info(f"✅ Produto cadastrado: {dados.get('nome_prod')}")
            return redirect(url_for('produto'))
            
        # PUT: Atualizar produto existente
        elif request.method == 'PUT' and id_prod:
            dados = request.form
            imagem = request.files.get('imagem_url')
            
            if imagem and imagem.filename:
                imagem_bytes = imagem.read()
                cur.execute("""
                    UPDATE tbl_prod 
                    SET nome_prod = %s, descricao = %s, subgrupo = %s, status_promocao = %s, 
                        valor = %s, form_pgmto = %s, imagem_url = %s
                    WHERE id_prod = %s
                """, (
                    dados.get('nome_prod', ''),
                    dados.get('descricao', ''),
                    dados.get('subgrupo', ''),
                    dados.get('status_promocao', ''),
                    dados.get('valor', 0),
                    dados.get('form_pgmto', ''),
                    imagem_bytes,
                    id_prod
                ))
            else:
                cur.execute("""
                    UPDATE tbl_prod 
                    SET nome_prod = %s, descricao = %s, subgrupo = %s, status_promocao = %s, 
                        valor = %s, form_pgmto = %s
                    WHERE id_prod = %s
                """, (
                    dados.get('nome_prod', ''),
                    dados.get('descricao', ''),
                    dados.get('subgrupo', ''),
                    dados.get('status_promocao', ''),
                    dados.get('valor', 0),
                    dados.get('form_pgmto', ''),
                    id_prod
                ))
            
            mysql.connection.commit()
            logging.info(f"✅ Produto atualizado: {dados.get('nome_prod')}")
            return jsonify({"message": "Produto atualizado com sucesso"})

        # DELETE: Excluir produto (soft delete - marcar como inativo)
        elif request.method == 'DELETE' and id_prod:
            try:
                # Verificar se o produto existe e está ativo
                cur.execute("SELECT id_prod, nome_prod FROM tbl_prod WHERE id_prod = %s AND ativo = 1", (id_prod,))
                produto_existe = cur.fetchone()
                
                if not produto_existe:
                    return jsonify({"error": "Produto não encontrado ou já foi deletado"}), 404
                
                # Marcar como inativo (soft delete)
                cur.execute("UPDATE tbl_prod SET ativo = 0 WHERE id_prod = %s", (id_prod,))
                
                if cur.rowcount > 0:
                    mysql.connection.commit()
                    logging.info(f"✅ Produto desativado: {id_prod} - {produto_existe['nome_prod']}")
                    return jsonify({"message": "Produto excluído com sucesso"}), 200
                else:
                    mysql.connection.rollback()
                    return jsonify({"error": "Falha ao excluir o produto"}), 400
                    
            except Exception as delete_error:
                mysql.connection.rollback()
                logging.error(f"❌ Erro ao deletar produto {id_prod}: {delete_error}")
                return jsonify({"error": f"Erro ao excluir: {str(delete_error)}"}), 500
            
    except Exception as e:
        logging.error(f"❌ Erro na rota /produto: {e}")
        if mysql.connection:
            mysql.connection.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if cur:
            cur.close()

@app.route('/produto_excel', methods=['GET'])
def produto_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor (agora funcionando corretamente)
        cur = mysql.connection.cursor(DictCursor)
        logging.info("✅ Cursor criado")

        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        produtos = cur.fetchall()

        logging.info(f"📊 Total de produtos: {len(produtos)}")

        if not produtos:
            return jsonify({"error": "Nenhum produto encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Produtos')

        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)

        logging.info("✍️ Cabeçalhos escritos")

        # --- LOOP CORRIGIDO ---
        for idx, produto in enumerate(produtos):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = produto[col]

                # Tratamentos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    # Evita texto ilegível de BLOB (imagem)
                    dados_linha.append("[BINARY_DATA]")
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(produtos)} linhas escritas")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'produtos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()


@app.route('/teste_produtos', methods=['GET'])
def teste_produtos():
    cur = None
    try:
        cur = mysql.connection.cursor(DictCursor)
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC LIMIT 3")

        colunas = [i[0] for i in cur.description]
        produtos = cur.fetchall()

        return jsonify({
            "status": "success",
            "total": len(produtos),
            "colunas": colunas,
            "produtos": produtos
        })

    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "erro": str(e),
            "traceback": traceback.format_exc()
        }), 500

    finally:
        if cur:
            cur.close()
            
@app.route('/pesquisa', methods=['GET', 'POST'])
def pesquisa():
    print(f"\n{'='*50}")
    print(f"MÉTODO: {request.method}")
    print(f"{'='*50}\n")
    
    if request.method == "POST":
        print("✅ Entrou no POST")
        print(f"📝 Dados recebidos: {dict(request.form)}\n")
        
        try:
            print("🔌 Tentando conectar ao MySQL...")
            conn = mysql.connection
            cur = conn.cursor()
            print("✅ Conectado ao banco!\n")

            # Criação da tabela
            print("📋 Criando/verificando tabela...")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tbl_pesquisa_satisfacao (
                    id_pesquisa INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                    atendimento VARCHAR(50),
                    qualidade VARCHAR(50),
                    satisfacao VARCHAR(50),
                    rapidez VARCHAR(50),
                    localizacao VARCHAR(50),
                    experiencia VARCHAR(50),
                    facilidade VARCHAR(50),
                    variedade VARCHAR(50),
                    ambiente VARCHAR(50),
                    recomendacao VARCHAR(50),
                    comentarios TEXT,
                    data_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            print("✅ Tabela OK!\n")

            # Coleta dos dados
            campos = ['atendimento', 'qualidade', 'satisfacao', 'rapidez', 'localizacao',
                    'experiencia', 'facilidade', 'variedade', 'ambiente', 'recomendacao']
            
            valores = [request.form.get(c) for c in campos]
            comentarios = request.form.get('comentarios', '')
            
            print("📊 Valores a serem inseridos:")
            for campo, valor in zip(campos, valores):
                print(f"  - {campo}: {valor}")
            print(f"  - comentarios: {comentarios}\n")

            # Inserção no banco
            print("💾 Inserindo no banco...")
            cur.execute("""
                INSERT INTO tbl_pesquisa_satisfacao (
                    atendimento, qualidade, satisfacao, rapidez, localizacao,
                    experiencia, facilidade, variedade, ambiente, recomendacao, comentarios
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (*valores, comentarios))
            
            print(f"✅ Linhas afetadas: {cur.rowcount}")
            
            conn.commit()
            print("✅ COMMIT realizado com sucesso!")
            
            cur.close()
            conn.close()
            print("🔒 Conexão fechada\n")

            return render_template('pesquisa.html', mensagem='Pesquisa enviada com sucesso!')

        except Exception as e:
            print(f"\n❌ ERRO CAPTURADO: {str(e)}\n")
            import traceback
            traceback.print_exc()
            return render_template('pesquisa.html', erro=str(e))
    
    print("📄 Renderizando formulário (GET)\n")
    return render_template('pesquisa.html')

@app.route('/contato', methods=['GET', 'POST'])
def contato():
    if request.method == 'POST':
        try:
            nome = request.form.get('nome')
            email = request.form.get('email')
            mensagem = request.form.get('mensagem')
            
            # Validação básica
            if not nome or not email or not mensagem:
                return render_template('contato.html', 
                                    erro="Todos os campos são obrigatórios!")
            
            # Criar cursor
            cur = mysql.connection.cursor()
            
            # Inserir dados
            cur.execute("""
                INSERT INTO u799109175_cestas_present.tbl_fale_conosco 
                (nome, email, mensagem)
                VALUES (%s, %s, %s)
            """, (nome, email, mensagem))
            
            # ORDEM CORRETA: commit antes de fechar
            mysql.connection.commit()
            cur.close()
            
            logging.info(f"📩 Mensagem recebida de {nome} ({email})")
            
            return render_template('contato.html', 
                                sucesso="Mensagem enviada com sucesso!")
        
        except Exception as e:
            logging.error(f"❌ Erro ao processar contato: {e}")
            return render_template('contato.html', 
                                erro="Erro ao enviar mensagem. Tente novamente.")
    
    return render_template('contato.html')

@app.route('/ger_clientes', methods=['GET', 'POST'])
def ger_clientes():
    if request.method == 'GET':
        try:
            cur = mysql.connection.cursor()
            
            # Captura os parâmetros de data
            data_inicio = request.args.get('data_inicio')
            data_fim = request.args.get('data_fim')
            
            # Query base 
            query = "SELECT * FROM tbl_cliente"
            params = []
            
            # Adiciona filtros se as datas forem fornecidas
            if data_inicio and data_fim:
                query += " WHERE DATE(created_at) BETWEEN %s AND %s"
                params = [data_inicio, data_fim]
            elif data_inicio:
                query += " WHERE DATE(created_at) >= %s"
                params = [data_inicio]
            elif data_fim:
                query += " WHERE DATE(created_at) <= %s"
                params = [data_fim]
            
            query += " ORDER BY nome_cliente ASC"
            
            # Executa a query com ou sem parâmetros
            if params:
                cur.execute(query, params)
            else:
                cur.execute(query)
                
            clientes = cur.fetchall()
            cur.close()
            
            return render_template('ger_clientes.html', clientes=clientes)
            
        except Exception as e:
            logging.error(f"❌ Erro ao buscar clientes: {e}")
            return render_template('ger_clientes.html', clientes=[])
    
    return render_template('ger_clientes.html')


@app.route('/cliente_excel', methods=['GET'])
def cliente_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        cur = mysql.connection.cursor(DictCursor)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM tbl_cliente")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Cliente: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum cliente encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Clientes')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


@app.route('/pedidos', methods=['GET', 'POST'])
def pedidos():
    cur = None
    try:
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM tbl_prod WHERE ativo = 1 ORDER BY created_at DESC")
        produtos = cur.fetchall()
        cur.execute("SELECT * FROM tbl_cliente")
        clientes = cur.fetchall()
        return render_template('pedidos.html', produtos=produtos, clientes=clientes)
    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos para pedidos: {e}")
        return render_template('pedidos.html', produtos=[], clientes=[])
    finally:
        if cur:
            cur.close()

    return render_template('pedidos.html')


@app.route('/salvar_pedido', methods=['POST'])
def salvar_pedido():
    cur = None
    try:
        dados = request.get_json()
        
        if not dados or 'carrinho' not in dados:
            return jsonify({"status": "erro", "mensagem": "Dados inválidos"}), 400
        
        carrinho = dados.get('carrinho', [])
        id_cliente = dados.get('id_cliente')
        nome_cliente = dados.get('nome_cliente')
        telefone_cliente = dados.get('telefone_cliente')
        
        if not carrinho:
            return jsonify({"status": "erro", "mensagem": "Carrinho vazio"}), 400
        
        if not id_cliente:
            return jsonify({"status": "erro", "mensagem": "ID do cliente obrigatório"}), 400
        
        cur = mysql.connection.cursor()
        
        # Não criar tabelas aqui — já existem no banco com a estrutura correta
        # O CREATE TABLE IF NOT EXISTS pode causar erro se a tabela foi alterada manualmente
        
        # Calcular valor total do pedido
        valor_total = sum(float(item.get('subtotal', 0)) for item in carrinho)
        
        # Inserir pedido principal
        cur.execute("""
            INSERT INTO tbl_pedidos (id_cliente, valor_total)
            VALUES (%s, %s)
        """, (id_cliente, valor_total))
        
        id_pedido = cur.lastrowid
        logging.info(f"✅ Pedido criado: {id_pedido}")
        
        # Inserir detalhes do pedido para cada item no carrinho
        for item in carrinho:
            id_prod = item.get('produtoId')
            quantidade = item.get('quantidade')
            preco_unitario = item.get('valor')
            valor_item = float(item.get('subtotal', 0))  # Subtotal do item
            
            cur.execute("""
                INSERT INTO tbl_detalhes_pedido 
                (id_pedido, id_prod, id_cliente, quantidade, preco_unitario, nome_cliente, telefone, valor_total)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (id_pedido, id_prod, id_cliente, quantidade, preco_unitario, nome_cliente, telefone_cliente, valor_item))
        
        mysql.connection.commit()
        logging.info(f"✅ Pedido salvo: {id_pedido} com {len(carrinho)} itens")
        
        return jsonify({
            "status": "sucesso",
            "mensagem": "Pedido salvo com sucesso!",
            "id_pedido": id_pedido,
            "valor_total": float(valor_total)
        }), 200
        
    except Exception as e:
        logging.error(f"❌ Erro ao salvar pedido: {e}")
        if mysql.connection:
            mysql.connection.rollback()
        return jsonify({"status": "erro", "mensagem": str(e)}), 500
    finally:
        if cur:
            cur.close()

#criar uma rota para enviar os pedidos via whatsapp
@app.route('/enviar_whatsapp', methods=['POST'])
def enviar_whatsapp():
    """
    Recebe dados do pedido e envia via WhatsApp usando Selenium
    """
    try:
        dados = request.get_json()
        
        if not dados:
            return jsonify({
                "status": "erro",
                "mensagem": "Dados não fornecidos"
            }), 400
        
        whatsapp_numero = dados.get('whatsapp_numero')
        mensagem = dados.get('mensagem')
        id_pedido = dados.get('id_pedido')
        
        if not whatsapp_numero or not mensagem:
            return jsonify({
                "status": "erro",
                "mensagem": "Número de WhatsApp ou mensagem não fornecidos"
            }), 400
        
        # Limpar número (remover caracteres especiais)
        whatsapp_numero = ''.join(filter(str.isdigit, whatsapp_numero))
        
        # Configurar Selenium
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Inicializar driver do Selenium
        driver = None
        try:
            driver = webdriver.Chrome(options=chrome_options)
            
            # URL do WhatsApp Web com mensagem pre-formatada
            url_whatsapp = f"https://wa.me/{whatsapp_numero}?text={quote(mensagem)}"
            
            logging.info(f"📱 Abrindo WhatsApp Web: {url_whatsapp}")
            driver.get(url_whatsapp)
            
            # Aguardar carregamento da página
            time.sleep(5)
            
            # Tentar encontrar o campo de mensagem
            try:
                # Esperar pelo input de mensagem aparecer
                message_input = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                )
                
                logging.info("✅ Campo de mensagem encontrado")
                
                # Esperar pelo botão de enviar
                send_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Enviar']"))
                )
                
                # Clicar no botão de enviar
                send_button.click()
                logging.info(f"✅ Mensagem enviada para {whatsapp_numero}")
                
                # Aguardar um pouco para garantir o envio
                time.sleep(3)
                
                # Atualizar status do pedido no banco de dados
                cur = mysql.connection.cursor()
                cur.execute("""
                    UPDATE tbl_pedidos 
                    SET status_pedido = 'enviado_whatsapp'
                    WHERE id_pedido = %s
                """, (id_pedido,))
                mysql.connection.commit()
                cur.close()
                
                logging.info(f"✅ Pedido {id_pedido} marcado como enviado via WhatsApp")
                
                return jsonify({
                    "status": "sucesso",
                    "mensagem": f"Pedido #{id_pedido} enviado via WhatsApp com sucesso!",
                    "id_pedido": id_pedido,
                    "numero_whatsapp": whatsapp_numero
                }), 200
                
            except Exception as e:
                logging.warning(f"⚠️ Erro ao enviar mensagem via Selenium: {e}")
                # Mesmo se falhar o envio automático, registrar no banco
                try:
                    cur = mysql.connection.cursor()
                    cur.execute("""
                        UPDATE tbl_pedidos 
                        SET status_pedido = 'pendente_whatsapp'
                        WHERE id_pedido = %s
                    """, (id_pedido,))
                    mysql.connection.commit()
                    cur.close()
                except:
                    pass
                
                raise Exception(f"Erro ao enviar mensagem: {str(e)}")
        
        finally:
            # Fechar o navegador
            if driver:
                driver.quit()
                logging.info("🔒 Navegador fechado")
    
    except Exception as e:
        logging.error(f"❌ Erro ao enviar WhatsApp: {e}")
        return jsonify({
            "status": "erro",
            "mensagem": f"Erro ao enviar pedido via WhatsApp: {str(e)}"
        }), 500


@app.route('/ger_pedidos', methods=['GET', 'POST'])
def ger_pedidos():
    if request.method == 'GET':
        try:
            cur = mysql.connection.cursor()
            
            # Captura os parâmetros de data
            data_inicio = request.args.get('data_inicio')
            data_fim = request.args.get('data_fim')
            
            # Query base 
            query = "SELECT * FROM vw_pedidos_fin"
            params = []
            
            # Adiciona filtros se as datas forem fornecidas
            if data_inicio and data_fim:
                query += " WHERE DATE(dt_registro) BETWEEN %s AND %s"
                params = [data_inicio, data_fim]
            elif data_inicio:
                query += " WHERE DATE(dt_registro) >= %s"
                params = [data_inicio]
            elif data_fim:
                query += " WHERE DATE(dt_registro) <= %s"
                params = [data_fim]
            
            query += " ORDER BY dt_registro DESC"
            
            # Executa a query com ou sem parâmetros
            if params:
                cur.execute(query, params)
            else:
                cur.execute(query)
                
            pedidos = cur.fetchall()
            cur.close()

            total_valor = sum([p['valor_total'] for p in pedidos])
            return render_template('ger_pedidos.html', pedidos=pedidos, total_valor=total_valor)
            
        except Exception as e:
            logging.error(f"❌ Erro ao buscar pedidos: {e}")
            return render_template('ger_pedidos.html', pedidos=[])
    
    return render_template('ger_pedidos.html')


@app.route('/pedidos_excel', methods=['GET'])
def pedidos_excel():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        cur = mysql.connection.cursor(DictCursor)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM vw_pedidos_fin")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Pedidos: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum pedido encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Pedidos')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


#=============================================================

@app.route('/pedidos_excel_clientes', methods=['GET'])
def pedidos_excel_clientes():
    cur = None
    try:
        logging.info("🚀 Iniciando geração do Excel...")

        # Cursor com DictCursor
        cur = mysql.connection.cursor(DictCursor)
        logging.info("✅ Cursor criado")

        # Buscar clientes
        cur.execute("SELECT * FROM vw_resumo_pedidos_cliente")
        logging.info("✅ Query executada")

        colunas = [i[0] for i in cur.description]
        dados = cur.fetchall()   # <- CORRIGIDO (não sobrescreve depois)

        logging.info(f"📊 Total de Pedidos: {len(dados)}")

        if not dados:
            return jsonify({"error": "Nenhum pedido encontrado"}), 404

        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Pedidos')

        # Formato do cabeçalho
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })

        # Formato padrão das células
        cell_format = workbook.add_format({'border': 1})

        # Escrever cabeçalhos
        worksheet.write_row(0, 0, colunas, header_format)
        logging.info("✍️ Cabeçalhos escritos")

        # LOOP sem sobrescrever a variável "dados"
        for idx, item in enumerate(dados):
            linha = idx + 1
            dados_linha = []

            for col in colunas:
                valor = item[col]

                # Tratamento de tipos
                if isinstance(valor, (datetime, date)):
                    dados_linha.append(valor.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(valor, bytes):
                    dados_linha.append("[BINARY_DATA]")  # evita bagunça de dados binários
                elif valor is None:
                    dados_linha.append('')
                else:
                    dados_linha.append(valor)

            worksheet.write_row(linha, 0, dados_linha, cell_format)

        logging.info(f"✅ {len(dados)} linhas escritas no Excel")

        # Ajuste das colunas
        for i in range(len(colunas)):
            worksheet.set_column(i, i, 20)

        worksheet.freeze_panes(1, 0)
        workbook.close()
        output.seek(0)

        logging.info("🎉 Excel gerado com sucesso!")

        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pedidos_cliente_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        erro_completo = traceback.format_exc()
        logging.error(f"❌ ERRO:\n{erro_completo}")

        return jsonify({
            "erro": str(e),
            "detalhes": erro_completo
        }), 500

    finally:
        if cur:
            cur.close()

        logging.info("✅ Cursor fechado")


@app.route('/subgrupo/<string:subgrupo>', methods=['GET'])
def grupo(subgrupo):
    cur = mysql.connection.cursor(dictionary=True)
    try:
        # Buscar produtos do subgrupo selecionado
        cur.execute("SELECT * FROM tbl_prod WHERE subgrupo = %s AND ativo = 1 ORDER BY created_at DESC", (subgrupo,))
        produtos = cur.fetchall()
        
        # Buscar todos os subgrupos para o filtro
        cur.execute("SELECT DISTINCT subgrupo FROM tbl_prod ORDER BY subgrupo ASC")
        subgrupos_raw = cur.fetchall()
        subgrupos = [{"subgrupo": s['subgrupo']} for s in subgrupos_raw] if subgrupos_raw else []

    except Exception as e:
        logging.error(f"❌ Erro ao buscar produtos do grupo: {e}")
        produtos = []
        subgrupos = []

    finally:
        if cur:
            cur.close()

    return render_template('index.html', produtos=produtos, subgrupos=subgrupos, subgrupo_selecionado=subgrupo)
